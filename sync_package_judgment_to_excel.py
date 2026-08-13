import pandas as pd
from openpyxl import Workbook
from pathlib import Path
import re
import unicodedata


# ============================================================
# ファイル設定
# ============================================================

# 正解として使用するCSV
REFERENCE_CSV = Path(
    "package_verification.csv"
)

# 検証・修正するCSV
TARGET_CSV = Path(
    "otc_master_updates_abuse_prevention.csv"
)

# 修正版CSV
OUTPUT_CSV = Path(
    "otc_master_updates_abuse_prevention_corrected.csv"
)

# 修正版Excel
OUTPUT_EXCEL = Path(
    "otc_master_updates_abuse_prevention_corrected.xlsx"
)


# ============================================================
# 文字コード
# ============================================================

# package_verification.csv
REFERENCE_ENCODING = "utf-8-sig"

# otc_master_updates_abuse_prevention.csv
TARGET_ENCODING = "cp932"


# ============================================================
# 検証対象CSVの列
# ============================================================

PRODUCT_COL = "商品名"

JUDGMENT_COL = "指定濫用防止容量"


# ============================================================
# 文字列を比較しやすくする
# ============================================================

def normalize_text(value):

    if value is None:
        return ""

    text = str(value)

    # 全角・半角を統一
    text = unicodedata.normalize("NFKC", text)

    # 空白を除去
    text = re.sub(r"\s+", "", text)

    # 先頭に付く可能性がある記号を除去
    text = text.replace("☆", "")
    text = text.replace("★", "")

    return text.strip()


# ============================================================
# 1. 正解CSVを読み込む
# ============================================================

print("")
print("==========================================")
print("1. 正解CSVを読み込みます")
print("==========================================")

reference_df = pd.read_csv(
    REFERENCE_CSV,
    encoding=REFERENCE_ENCODING
)

print(
    f"正解CSV：{REFERENCE_CSV}"
)

print(
    f"件数：{len(reference_df)}"
)


# ============================================================
# 正解CSVの列を確認
# ============================================================

required_reference_columns = [
    "product",
    "package",
    "judgment"
]

missing_reference = [
    col
    for col in required_reference_columns
    if col not in reference_df.columns
]

if missing_reference:

    print("")
    print("【エラー】")
    print("package_verification.csv に必要な列がありません。")

    print("不足している列：")

    for col in missing_reference:
        print(f"  - {col}")

    print("")
    print("実際の列：")
    print(list(reference_df.columns))

    raise SystemExit(1)


# ============================================================
# 2. 正解データを辞書化
# ============================================================

judgment_map = {}

for _, row in reference_df.iterrows():

    product = normalize_text(
        row["product"]
    )

    package = normalize_text(
        row["package"]
    )

    judgment = str(
        row["judgment"]
    ).strip()

    if not product or not package:
        continue

    key = (
        product,
        package
    )

    judgment_map[key] = judgment


print(
    f"正解判定：{len(judgment_map)}件"
)


# ============================================================
# 3. 検証対象CSVを読み込む
# ============================================================

print("")
print("==========================================")
print("2. 検証対象CSVを読み込みます")
print("==========================================")

target_df = pd.read_csv(
    TARGET_CSV,
    encoding="cp932"
)

print(
    f"検証対象CSV：{TARGET_CSV}"
)

print(
    f"件数：{len(target_df)}"
)


# ============================================================
# 4. 検証対象CSVの列を確認
# ============================================================

required_target_columns = [
    PRODUCT_COL,
    JUDGMENT_COL
]

missing_target = [
    col
    for col in required_target_columns
    if col not in target_df.columns
]

if missing_target:

    print("")
    print("【エラー】")
    print("検証対象CSVに必要な列がありません。")

    print("不足している列：")

    for col in missing_target:
        print(f"  - {col}")

    print("")
    print("実際の列：")
    print(list(target_df.columns))

    raise SystemExit(1)


print("")
print("検証対象CSVの列：")

for col in target_df.columns:
    print(f"  - {col}")


# ============================================================
# 5. 商品名を照合して大小判定を修正
# ============================================================

print("")
print("==========================================")
print("3. 正解CSVと照合します")
print("==========================================")

updated_count = 0
unchanged_count = 0
not_found_count = 0

changes = []
not_found = []


for index, row in target_df.iterrows():

    product = normalize_text(
        row[PRODUCT_COL]
    )

    if not product:
        continue

    current_judgment = str(
        row[JUDGMENT_COL]
    ).strip()


    # --------------------------------------------------------
    # 商品名から正解CSVの候補を探す
    # --------------------------------------------------------

    candidates = []

    for (
        ref_product,
        ref_package
    ), judgment in judgment_map.items():

        # 商品名が一致・包含関係にあるものを候補にする
        if (
            ref_product in product
            or product in ref_product
        ):

            candidates.append(
                (
                    ref_product,
                    ref_package,
                    judgment
                )
            )


    # --------------------------------------------------------
    # 候補がない
    # --------------------------------------------------------

    if not candidates:

        not_found_count += 1

        not_found.append(
            {
                "row": index + 2,
                "product": product
            }
        )

        continue


    # --------------------------------------------------------
    # 候補が複数ある場合
    # --------------------------------------------------------

    # 商品名との一致が最も長い候補を優先
    candidates.sort(
        key=lambda x: len(x[0]),
        reverse=True
    )

    best_product = candidates[0][0]

    best_candidates = [
        c
        for c in candidates
        if c[0] == best_product
    ]


    # 同じ商品名で複数包装が存在する場合
    # 商品名に包装情報が含まれているかを確認
    matched = None

    for candidate in best_candidates:

        ref_product = candidate[0]
        ref_package = normalize_text(
            candidate[1]
        )

        judgment = candidate[2]

        # 商品名そのものに包装表記が含まれている場合
        if ref_package in product:

            matched = judgment
            break


    # 一意に判断できなかった場合
    if matched is None:

        if len(best_candidates) == 1:

            matched = best_candidates[0][2]

        else:

            not_found_count += 1

            not_found.append(
                {
                    "row": index + 2,
                    "product": product
                }
            )

            continue


    correct_judgment = str(
        matched
    ).strip()


    # --------------------------------------------------------
    # 判定が違えば修正
    # --------------------------------------------------------

    if current_judgment != correct_judgment:

        target_df.at[
            index,
            JUDGMENT_COL
        ] = correct_judgment

        updated_count += 1

        changes.append(
            {
                "row": index + 2,
                "product": product,
                "before": current_judgment,
                "after": correct_judgment
            }
        )

        print(
            f"修正：{product}"
        )

        print(
            f"  {current_judgment} → {correct_judgment}"
        )

    else:

        unchanged_count += 1


# ============================================================
# 6. 修正版CSVを保存
# ============================================================

target_df.to_csv(
    OUTPUT_CSV,
    index=False,
    encoding="utf-8-sig"
)


# ============================================================
# 7. 修正版Excelを作成
# ============================================================

print("")
print("==========================================")
print("4. 修正版Excelを作成します")
print("==========================================")

wb = Workbook()

ws = wb.active

ws.title = "修正版"


# ヘッダー

for col_index, column_name in enumerate(
    target_df.columns,
    start=1
):

    ws.cell(
        row=1,
        column=col_index
    ).value = column_name


# データ

for row_index, row in enumerate(
    target_df.itertuples(index=False),
    start=2
):

    for col_index, value in enumerate(
        row,
        start=1
    ):

        ws.cell(
            row=row_index,
            column=col_index
        ).value = value


# 列幅を自動調整

for column_cells in ws.columns:

    max_length = 0

    column_letter = column_cells[0].column_letter

    for cell in column_cells:

        if cell.value is not None:

            length = len(
                str(cell.value)
            )

            max_length = max(
                max_length,
                length
            )

    ws.column_dimensions[
        column_letter
    ].width = min(
        max_length + 2,
        60
    )


# 保存

wb.save(
    OUTPUT_EXCEL
)


# ============================================================
# 8. 結果表示
# ============================================================

print("")
print("==========================================")
print("完了")
print("==========================================")

print(
    f"修正件数：{updated_count}件"
)

print(
    f"変更なし：{unchanged_count}件"
)

print(
    f"照合できなかった件数：{not_found_count}件"
)

print("")
print(
    f"修正版CSV：{OUTPUT_CSV}"
)

print(
    f"修正版Excel：{OUTPUT_EXCEL}"
)


# ============================================================
# 9. 照合できなかった商品
# ============================================================

if not_found:

    print("")
    print("==========================================")
    print("照合できなかった商品")
    print("==========================================")

    for item in not_found[:30]:

        print(
            f"行{item['row']}："
            f"{item['product']}"
        )

    if len(not_found) > 30:

        print(
            f"... 他 {len(not_found) - 30}件"
        )


# ============================================================
# 10. 修正内容
# ============================================================

if changes:

    print("")
    print("==========================================")
    print("修正内容")
    print("==========================================")

    for item in changes[:50]:

        print(
            f"{item['product']} : "
            f"{item['before']} → "
            f"{item['after']}"
        )

    if len(changes) > 50:

        print(
            f"... 他 {len(changes) - 50}件"
        )